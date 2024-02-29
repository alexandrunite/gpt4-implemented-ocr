from openai import OpenAI
from openpyxl import load_workbook
import smtplib
from pdf2image import convert_from_path
from string import Template
import requests
from datetime import datetime
import base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import json
import os
from openpyxl import load_workbook
#import magic
from random import randint
from time import sleep

numar=0
MY_ADDRESS = ''
PASSWORD = ''
email=''
excel_path=r""
folder_path=r""
api_key = ""
image_path = r""
image_path_2 = rf""

def scan_for_files():
    files=os.listdir(folder_path)
    new_files=[file for file in files if os.path.isfile(os.path.join(folder_path, file))]
    if new_files:
        return True
    else:
        return False

def is_pdf_file(file_path):
    mime_type = magic.from_file(file_path, mime=True)
    return mime_type == 'application/pdf'

def conversie_pdf():
    files = os.listdir(folder_path)
    new_files=[file for file in files if os.path.isfile(os.path.join(folder_path, file))]
    pdf_path = ''
    images = convert_from_path(pdf_path)
    num=0
    for i, image in enumerate(images):
        num+=1
        image.save(f'page_{i+1}.jpeg', 'JPEG')
    numar=num


#merge
def encode_image(image_path):
  with open(image_path, "rb") as image_file:
    return base64.b64encode(image_file.read()).decode('utf-8')

#merge
def datetime_generation():
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%y, ora %H:%M:%S")
    ls=formatted_datetime.split()
    ls_new=[]
    text_final=" ".join(ls_new)
    return text_final

def delete_file(file_path):
    try:
        os.remove(file_path)
        print(f"The file {file_path} has been deleted successfully.")
    except OSError as e:
        print(f"Error occurred while deleting the file: {e}")

#merge
def message_generation():
    message=f""
    attachment=r""
    return message, attachment

#merge
def send_mail():
    mesaj,atasament=message_generation()
    msg= MIMEMultipart()
    msg["From"]=MY_ADDRESS
    msg["To"]=email
    msg["Subject"]=f"Excel Facturi {datetime_generation()}"
    msg.attach(MIMEText(mesaj,"plain"))
    atasament_deschis=open(atasament,"rb")
    part=MIMEBase("application","octet-stream")
    part.set_payload((atasament_deschis).read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename=excelfacturi.xlsx")
    msg.attach(part)
    text = msg.as_string()
    smtp_server="smtp.gmail.com"
    smtp_port=465
    with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
        server.login(MY_ADDRESS, PASSWORD)
        server.sendmail(MY_ADDRESS, email, text)

#merge
def data_scraper():
    base64_image = encode_image(image_path)
    headers = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {api_key}"
    }
    payload = {
    "model": "gpt-4-vision-preview",
    "messages": [
        {
        "role": "user",
        "content": [
            {
            "type": "text",
            "text": "You will analyze it and extract the following details: 1. The date when the invoice was issued, formatted as day.month.year. 2. The name of the store or provider (furnizor) that issued the invoice. 3. The total sum listed on the invoice without VAT (value-added tax, also known as TVA). 4. The amount of VAT (TVA) charged, if applicable. 5. The total sum including VAT (TVA). The extracted information will be presented PRECISELY in the following format: \"day.month.year, name of furnizor/store, totalSumWithoutVAT, VAT, totalSumWithVAT\", i don't wany any extra text other than the presented format"
            },
            {
            "type": "image_url",
            "image_url": {
                "url": f"data:image/jpeg;base64,{base64_image}"
            }
            }
        ]
        }
    ],
    "max_tokens": 300
    }
    response = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload)
    print(response.json())
    datele=response.json()
    data,furnizor,totalFaraTVA,taxaTVA,totalCuTVA=datele['choices'][0]['message']['content'].split(", ")
    return data,furnizor,totalFaraTVA,taxaTVA,totalCuTVA

#merge
def deschidere_adaugare():
    workbook = load_workbook(filename=excel_path)
    sheet = workbook.active
    data, furnizor, totalFaraTVA, taxaTVA, totalCuTVA = data_scraper()
    row_number = sheet.max_row +1
    print(row_number)
    data_to_insert = [row_number, furnizor, data, totalFaraTVA, taxaTVA, totalCuTVA]
    sheet.insert_rows(row_number)
    for col_idx, data in enumerate(data_to_insert, start=1):
        cell = sheet.cell(row=row_number, column=col_idx)
        cell.value = data
    workbook.save(filename=excel_path)

deschidere_adaugare()